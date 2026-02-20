# -*- coding: utf-8 -*-
"""
Export Service - Report Generation

Handles generation of PDF and Excel reports.
"""

import io
from typing import Dict, Any, Optional
from datetime import datetime


class ExportService:
    """
    Service for generating export reports.
    
    Supports PDF and Excel formats.
    """
    
    def __init__(self):
        """Initialize export service."""
        pass
    
    def generate_excel_report(
        self,
        data: Dict[str, Any],
        include_charts: bool = True
    ) -> bytes:
        """
        Generate Excel report from analytics data.
        
        Args:
            data: Analytics data dictionary.
            include_charts: Whether to include chart data.
        
        Returns:
            bytes: Excel file content.
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("Please install openpyxl: pip install openpyxl")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==========================================================================
        # Summary Sheet
        # ==========================================================================
        ws_summary = wb.create_sheet('Summary')
        ws_summary.append(['ShopSense AI - Analytics Report'])
        ws_summary.append(['Generated:', datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')])
        ws_summary.append([])
        
        # KPIs
        if 'kpis' in data:
            kpis = data['kpis']
            ws_summary.append(['Key Metrics'])
            ws_summary.append(['Metric', 'Value'])
            
            metrics = [
                ('Total Revenue', f"${kpis.get('total_revenue', 0):,.2f}"),
                ('Total Units', f"{kpis.get('total_units', 0):,}"),
                ('Total Products', f"{kpis.get('total_products', 0):,}"),
                ('Avg Order Value', f"${kpis.get('avg_order_value', 0):,.2f}"),
                ('Avg Price', f"${kpis.get('avg_price', 0):,.2f}")
            ]
            
            for metric in metrics:
                ws_summary.append(metric)
        
        ws_summary.append([])
        
        # Recommendations
        if 'analysis' in data and 'recommendations' in data['analysis']:
            ws_summary.append(['Recommendations'])
            ws_summary.append(['Priority', 'Category', 'Recommendation'])
            
            for rec in data['analysis']['recommendations']:
                ws_summary.append([
                    rec.get('priority', '').upper(),
                    rec.get('category', ''),
                    rec.get('recommendation', '')
                ])
        
        # Format summary sheet
        self._format_sheet(ws_summary, header_font, header_fill, header_alignment, thin_border)
        
        # ==========================================================================
        # Products Sheet
        # ==========================================================================
        if 'charts' in data and 'top_products' in data['charts']:
            ws_products = wb.create_sheet('Products')
            ws_products.append(['Product Performance'])
            ws_products.append([])
            ws_products.append(['Product Name', 'Units Sold', 'Price', 'Revenue'])
            
            for product in data['charts']['top_products']:
                ws_products.append([
                    product.get('product_name', ''),
                    product.get('units_sold', 0),
                    f"${product.get('price', 0):,.2f}",
                    f"${product.get('revenue', 0):,.2f}"
                ])
            
            self._format_sheet(ws_products, header_font, header_fill, header_alignment, thin_border)
        
        # ==========================================================================
        # Trends Sheet
        # ==========================================================================
        if 'charts' in data and 'time_series' in data['charts']:
            ws_trends = wb.create_sheet('Trends')
            ws_trends.append(['Sales Trends'])
            ws_trends.append([])
            ws_trends.append(['Date', 'Units Sold', 'Revenue'])
            
            for trend in data['charts']['time_series']:
                ws_trends.append([
                    trend.get('date', ''),
                    trend.get('units_sold', 0),
                    f"${trend.get('revenue', 0):,.2f}"
                ])
            
            self._format_sheet(ws_trends, header_font, header_fill, header_alignment, thin_border)
        
        # ==========================================================================
        # Forecast Sheet
        # ==========================================================================
        if 'charts' in data and 'forecast' in data['charts']:
            ws_forecast = wb.create_sheet('Forecast')
            ws_forecast.append(['Sales Forecast (30 Days)'])
            ws_forecast.append([])
            ws_forecast.append(['Date', 'Predicted Revenue', 'Lower Bound', 'Upper Bound'])
            
            for forecast in data['charts']['forecast']:
                ws_forecast.append([
                    forecast.get('date', ''),
                    f"${forecast.get('predicted_revenue', 0):,.2f}",
                    f"${forecast.get('lower_bound', 0):,.2f}" if forecast.get('lower_bound') else 'N/A',
                    f"${forecast.get('upper_bound', 0):,.2f}" if forecast.get('upper_bound') else 'N/A'
                ])
            
            self._format_sheet(ws_forecast, header_font, header_fill, header_alignment, thin_border)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
    
    def generate_csv_export(
        self,
        data: list,
        filename: str = 'export.csv'
    ) -> bytes:
        """
        Generate CSV export from data.
        
        Args:
            data: List of dictionaries.
            filename: Output filename.
        
        Returns:
            bytes: CSV file content.
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("Please install pandas: pip install pandas")
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        
        return output.read()
    
    def _format_sheet(self, ws, header_font, header_fill, header_alignment, thin_border):
        """
        Apply formatting to worksheet.
        
        Args:
            ws: OpenPyXL worksheet.
            header_font: Font style for headers.
            header_fill: Fill style for headers.
            header_alignment: Alignment for headers.
            thin_border: Border style for cells.
        """
        # Format headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


# Singleton instance
export_service = ExportService()
